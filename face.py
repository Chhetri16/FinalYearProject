import cv2
import face_recognition
import os
import pickle
from tkinter import Tk, simpledialog, messagebox
import time
from openpyxl import Workbook, load_workbook

# Function to load face encodings from a file
def load_encodings():
    if os.path.exists('encodings.pickle'):
        with open('encodings.pickle', 'rb') as f:
            encodings = pickle.load(f)
        return encodings
    else:
        return {}

# Function to save face encodings to a file
def save_encodings(encodings):
    with open('encodings.pickle', 'wb') as f:
        pickle.dump(encodings, f)

# Function to get user input using Tkinter
def get_user_input():
    return simpledialog.askstring("Input", "Enter Name:")

# Function to get both name and roll number from the user
def get_name_and_roll():
    name = simpledialog.askstring("Input", "Enter Name:")
    roll = simpledialog.askstring("Input", "Enter Roll No:")
    return name, roll

# Load existing face encodings
face_encodings = load_encodings()

# Set to store names of faces for which attendance has been logged
attendance_logged_faces = set()

# Get the file name using Tkinter
root = Tk()
root.withdraw()  # Hide the main window

fileName = simpledialog.askstring("Input", "Enter Date:")

# Create an Excel workbook or load existing workbook
if os.path.exists(f'{fileName}_attendance.xlsx'):
    wb = load_workbook(f'{fileName}_attendance.xlsx')
else:
    wb = Workbook()

# Select active worksheet
ws = wb.active
ws.append(["Name", "Roll No."])  # Adding header if it's a new file
# Open a video capture object (0 for the default camera)
cap = cv2.VideoCapture(0)
cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)  # Set width
cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480) # Set height

# Allow the camera to adjust by adding a delay
time.sleep(2)

# Roll number variable declaration outside the loop
roll = None

while True:
    # Read a frame from the video capture
    ret, frame = cap.read()

    # Check if the camera is opened successfully
    if not ret:
        print("Error: Failed to open camera")
        break

    # Find face locations and encodings in the current frame
    face_locations = face_recognition.face_locations(frame)
    current_face_encodings = face_recognition.face_encodings(frame, face_locations)

    for face_encoding, face_location in zip(current_face_encodings, face_locations):
        # Check if the face is already known
        matches = face_recognition.compare_faces(list(face_encodings.values()), face_encoding)
        name = "Unknown"

        # If a match is found, use the name of the known face
        if True in matches:
            first_match_index = matches.index(True)
            name = list(face_encodings.keys())[first_match_index]
        else:
            # If the face is unknown, prompt the user to enter name and roll number
            name, roll = get_name_and_roll()  # Corrected line
            face_encodings[name] = face_encoding
            save_encodings(face_encodings)
            # Write name and roll number to Excel file
            ws.append([name, roll])
            wb.save(f'{fileName}_attendance.xlsx')

        # Draw a rectangle around the face and display the name
        top, right, bottom, left = face_location
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

        # Concatenate name and roll for display
        text = f"{name} ({roll})" if roll is not None else name

        cv2.putText(frame, text, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)  # Corrected line

        # Log attendance if the face is not in the set
        if name not in attendance_logged_faces:
            with open(f'{fileName}_attendance.txt', 'a') as log_file:
                log_file.write(f"{name} ({roll})\n")
            attendance_logged_faces.add(name)

    # Display the frame with face detection
    cv2.imshow('Camera', frame)

    # Break the loop when 'q' is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the video capture object and close the window
cap.release()
cv2.destroyAllWindows()
